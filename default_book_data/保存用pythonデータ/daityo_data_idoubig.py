import openpyxl
import json

file_name = "./data/図書台帳(印刷用) - 修正版.xlsx" #エクセルファイルの配置を設定

wb = openpyxl.load_workbook(file_name)

sheet = wb['分類表 (新)']

mid_list = []

cell1 = 0
cell2 = ""
cell3 = 0
cell4 = ""
num = 2
while True:
  if sheet.cell(row = num,column = 1).value is not None:
    data1 = sheet.cell(row = num,column = 1).value
    if  str == type(data1):
      data1 = int(data1)
  if sheet.cell(row = num,column = 2).value is not None:
    data2 = sheet.cell(row = num,column = 2).value
  if sheet.cell(row = num,column = 3).value is not None:
    data3 = sheet.cell(row = num,column = 3).value
    if  str == type(data3):
      data3 = int(data3)
  data4 = sheet.cell(row = num,column = 4).value

  if data4 is None:
    break
  
  cell_data={"大分類番号":data1,"大分類名":data2,"中分類番号":data3,"中分類名":data4}

  mid_list.append(cell_data)

  num += 1

js = open('./data/catecory_list.json','w',encoding="utf-8")#データを吐き出す場所を設定

json.dump(mid_list,js,indent=4, ensure_ascii=False)